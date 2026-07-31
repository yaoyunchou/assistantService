"""发布成功后回填汇总表与单品 Excel。

只写目标单元格、不整表重建，避免丢掉用户在表里做的格式与公式。
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from spider.goofish.config import (
    GOOFISH_DATA_DIR,
    ITEM_URL_TEMPLATE,
    PRODUCT_EXCEL_NAME,
    SUMMARY_EXCEL_NAME,
)
from spider.goofish.data.loader import (
    COL_ITEM_ID,
    COL_TITLE,
    COL_UPLOAD_LINK,
    COL_UPLOAD_TIME,
)
from utils.logger import get_logger

logger = get_logger('GoofishBackfill')


def _safe_dirname(title: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', title or '').strip()[:80]


def _resolve_summary_excel(root: Path) -> Path:
    named = root / SUMMARY_EXCEL_NAME
    if named.exists():
        return named
    for p in sorted(root.glob('*.xlsx')):
        if p.name.startswith('~$') or '.tmp' in p.name.lower():
            continue
        if '汇总' in p.name:
            return p
    return named


def build_item_url(item_id: str) -> str:
    item_id = str(item_id or '').strip()
    if not item_id:
        return ''
    return ITEM_URL_TEMPLATE.format(item_id=item_id)


def backfill_upload_result(
    *,
    title: str,
    item_id: str,
    item_url: str = '',
    product_dir: Optional[Path] = None,
    root: Optional[Path] = None,
) -> dict:
    """回填上架时间/链接/商品ID。

    Returns:
        { ok, item_url, upload_time, summary_updated, product_updated }
    """
    root = root or GOOFISH_DATA_DIR
    upload_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    url = item_url or build_item_url(item_id)
    product_dir = product_dir or (root / _safe_dirname(title))

    summary_updated = _backfill_summary(title, item_id, url, upload_time, root)
    product_updated = _backfill_product_excel(product_dir, item_id, url, upload_time)

    logger.info(
        '回填完成 title=%s item_id=%s summary=%s product=%s',
        title, item_id, summary_updated, product_updated,
    )
    return {
        'ok': True,
        'item_url': url,
        'upload_time': upload_time,
        'summary_updated': summary_updated,
        'product_updated': product_updated,
    }


def _backfill_summary(title: str, item_id: str, item_url: str, upload_time: str, root: Path) -> bool:
    summary_path = _resolve_summary_excel(root)
    if not summary_path.exists():
        logger.warning('汇总表不存在，跳过回填: %s', summary_path)
        return False

    wb = load_workbook(str(summary_path))
    try:
        ws = wb.active
        headers = [c.value for c in ws[1]]

        def col_of(name: str) -> Optional[int]:
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        col_title = col_of(COL_TITLE)
        if col_title is None:
            logger.warning('汇总表缺少「%s」列，无法定位行', COL_TITLE)
            return False

        col_time = col_of(COL_UPLOAD_TIME)
        col_link = col_of(COL_UPLOAD_LINK)
        col_item = col_of(COL_ITEM_ID)

        found = False
        for row in ws.iter_rows(min_row=2):
            cell_title = row[col_title - 1].value
            if cell_title is None or str(cell_title).strip() != str(title).strip():
                continue
            r = row[0].row
            if col_time:
                ws.cell(row=r, column=col_time, value=upload_time)
            if col_link:
                ws.cell(row=r, column=col_link, value=item_url)
            if col_item:
                ws.cell(row=r, column=col_item, value=item_id)
            found = True
            break

        if found:
            wb.save(str(summary_path))
        else:
            logger.warning('汇总表未找到标题为「%s」的行', title)
        return found
    finally:
        wb.close()


def _backfill_product_excel(product_dir: Path, item_id: str, item_url: str, upload_time: str) -> bool:
    excel_path = product_dir / PRODUCT_EXCEL_NAME
    if not excel_path.exists():
        return False

    wb = load_workbook(str(excel_path))
    try:
        if '基本信息' not in wb.sheetnames:
            return False
        ws = wb['基本信息']
        fields = {
            COL_UPLOAD_TIME: upload_time,
            COL_UPLOAD_LINK: item_url,
            COL_ITEM_ID: item_id,
        }
        existing = {
            ws.cell(row=r, column=1).value: r
            for r in range(2, ws.max_row + 1)
        }
        for name, value in fields.items():
            if name in existing:
                ws.cell(row=existing[name], column=2, value=value)
            else:
                next_row = ws.max_row + 1
                ws.cell(row=next_row, column=1, value=name)
                ws.cell(row=next_row, column=2, value=value)
        wb.save(str(excel_path))
        return True
    finally:
        wb.close()

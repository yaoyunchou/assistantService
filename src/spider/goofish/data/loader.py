"""从 Excel 与 images 目录加载待发布的闲鱼商品。

汇总表列头（缺列自动回落到缺省值，便于表结构逐步补全）：
    商品标题 | 想卖价 | 原价 | 成色 | 分类 | 是否包邮 | 发货地 | 图片数
    本地目录 | 商品URL | 采集时间 | 上架时间 | 上架链接 | 闲鱼商品ID | 状态
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from spider.goofish.config import (
    DEFAULT_CATEGORY,
    DEFAULT_CONDITION,
    DEFAULT_FREE_SHIPPING,
    DEFAULT_SHIP_FROM,
    GOOFISH_DATA_DIR,
    PRODUCT_EXCEL_NAME,
    SUMMARY_EXCEL_NAME,
)
from utils.logger import get_logger

logger = get_logger('GoofishLoader')

_IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}

COL_TITLE = '商品标题'
COL_PRICE = '想卖价'
COL_ORIGINAL_PRICE = '原价'
COL_CONDITION = '成色'
COL_CATEGORY = '分类'
COL_FREE_SHIPPING = '是否包邮'
COL_SHIP_FROM = '发货地'
COL_DESCRIPTION = '描述'
COL_DIR = '本地目录'
COL_PRODUCT_URL = '商品URL'
COL_UPLOAD_TIME = '上架时间'
COL_UPLOAD_LINK = '上架链接'
COL_ITEM_ID = '闲鱼商品ID'

SUMMARY_HEADERS = [
    COL_TITLE, COL_PRICE, COL_ORIGINAL_PRICE, COL_CONDITION, COL_CATEGORY,
    COL_FREE_SHIPPING, COL_SHIP_FROM, '图片数', COL_DIR, COL_PRODUCT_URL,
    '采集时间', COL_UPLOAD_TIME, COL_UPLOAD_LINK, COL_ITEM_ID, '状态',
]

_TRUTHY = {'是', 'true', '1', 'y', 'yes', '包邮', 'v'}
_FALSY = {'否', 'false', '0', 'n', 'no', '不包邮'}


def _to_bool(value: Any, default: bool) -> bool:
    text = str(value or '').strip().lower()
    if not text:
        return default
    if text in _TRUTHY:
        return True
    if text in _FALSY:
        return False
    return default


@dataclass
class ProductRecord:
    """闲鱼商品记录。price 为「想卖价」。"""

    title: str
    price: str
    product_dir: Path
    description: str = ''
    original_price: str = ''
    condition: str = DEFAULT_CONDITION
    category: str = DEFAULT_CATEGORY
    free_shipping: bool = DEFAULT_FREE_SHIPPING
    ship_from: str = DEFAULT_SHIP_FROM
    product_url: str = ''
    images: List[Path] = field(default_factory=list)
    upload_link: str = ''
    upload_time: str = ''
    item_id: str = ''

    @property
    def slug(self) -> str:
        illegal = r'[\\/:*?"<>|]'
        name = re.sub(illegal, '_', self.title).strip()
        return name[:40] or 'product'

    def to_dict(self) -> Dict[str, Any]:
        return {
            'title': self.title,
            'price': self.price,
            'original_price': self.original_price,
            'description': self.description,
            'condition': self.condition,
            'category': self.category,
            'free_shipping': self.free_shipping,
            'ship_from': self.ship_from,
            'product_dir': str(self.product_dir),
            'product_url': self.product_url,
            'images': [str(p) for p in self.images],
            'image_count': len(self.images),
            'upload_link': self.upload_link,
            'upload_time': self.upload_time,
            'item_id': self.item_id,
        }

    def missing_required(self) -> List[str]:
        """返回缺失的必填字段名（发布前校验）。"""
        missing = []
        if not str(self.title).strip():
            missing.append(COL_TITLE)
        if not str(self.price).strip():
            missing.append(COL_PRICE)
        if not self.images:
            missing.append('图片')
        return missing


def _cell_str(val: Any) -> str:
    if val is None:
        return ''
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _resolve_summary_excel(root: Path) -> Path:
    named = root / SUMMARY_EXCEL_NAME
    if named.exists():
        return named
    candidates = [
        p for p in sorted(root.glob('*.xlsx'))
        if not p.name.startswith('~$') and '.tmp' not in p.name.lower()
    ]
    for p in candidates:
        if '汇总' in p.name:
            return p
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f'未找到汇总 Excel: {named}')


def _collect_images(product_dir: Path) -> List[Path]:
    img_dir = product_dir / 'images'
    if not img_dir.is_dir():
        return []
    files = [
        p for p in img_dir.iterdir()
        if p.is_file() and p.suffix.lower() in _IMAGE_EXTS
    ]
    return sorted(files, key=lambda p: p.name.lower())


def _read_sheet_kv(ws) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out[_cell_str(row[0])] = _cell_str(row[1] if len(row) > 1 else '')
    return out


def _resolve_product_dir(row: Dict[str, str], root: Path) -> Optional[Path]:
    raw = row.get(COL_DIR) or ''
    if raw:
        candidate = Path(raw)
        if candidate.is_dir():
            return candidate
        if candidate.name:
            fallback = root / candidate.name
            if fallback.is_dir():
                return fallback
    title = row.get(COL_TITLE) or ''
    if title:
        guess = root / re.sub(r'[\\/:*?"<>|]', '_', title).strip()[:80]
        if guess.is_dir():
            return guess
    return None


def _load_product_detail(product_dir: Path, row: Dict[str, str]) -> ProductRecord:
    title = row.get(COL_TITLE) or product_dir.name
    price = row.get(COL_PRICE) or ''
    original_price = row.get(COL_ORIGINAL_PRICE) or ''
    condition = row.get(COL_CONDITION) or ''
    category = row.get(COL_CATEGORY) or ''
    ship_from = row.get(COL_SHIP_FROM) or ''
    description = row.get(COL_DESCRIPTION) or ''
    free_shipping_raw = row.get(COL_FREE_SHIPPING)
    product_url = row.get(COL_PRODUCT_URL) or ''
    upload_link = row.get(COL_UPLOAD_LINK) or ''
    upload_time = row.get(COL_UPLOAD_TIME) or ''
    item_id = row.get(COL_ITEM_ID) or ''

    excel_path = product_dir / PRODUCT_EXCEL_NAME
    if excel_path.exists():
        try:
            wb = load_workbook(str(excel_path), read_only=True, data_only=True)
            try:
                basic = _read_sheet_kv(wb['基本信息']) if '基本信息' in wb.sheetnames else {}
            finally:
                wb.close()
            title = basic.get(COL_TITLE) or title
            price = basic.get(COL_PRICE) or basic.get('价格（元）') or basic.get('价格') or price
            original_price = basic.get(COL_ORIGINAL_PRICE) or original_price
            condition = basic.get(COL_CONDITION) or condition
            category = basic.get(COL_CATEGORY) or category
            ship_from = basic.get(COL_SHIP_FROM) or ship_from
            description = basic.get(COL_DESCRIPTION) or description
            if basic.get(COL_FREE_SHIPPING):
                free_shipping_raw = basic.get(COL_FREE_SHIPPING)
            product_url = basic.get(COL_PRODUCT_URL) or product_url
            upload_link = basic.get(COL_UPLOAD_LINK) or upload_link
            upload_time = basic.get(COL_UPLOAD_TIME) or upload_time
            item_id = basic.get(COL_ITEM_ID) or item_id
        except Exception as exc:
            logger.warning('读取单品 Excel 失败 %s: %s', excel_path, exc)

    if not description:
        desc_txt = product_dir / '描述.txt'
        if desc_txt.exists():
            try:
                description = desc_txt.read_text(encoding='utf-8').strip()
            except Exception:
                pass

    return ProductRecord(
        title=title,
        price=price,
        product_dir=product_dir,
        description=description,
        original_price=original_price,
        condition=condition or DEFAULT_CONDITION,
        category=category or DEFAULT_CATEGORY,
        free_shipping=_to_bool(free_shipping_raw, DEFAULT_FREE_SHIPPING),
        ship_from=ship_from or DEFAULT_SHIP_FROM,
        product_url=product_url,
        images=_collect_images(product_dir),
        upload_link=upload_link,
        upload_time=upload_time,
        item_id=item_id,
    )


def _iter_summary_rows(root: Optional[Path] = None) -> List[Dict[str, str]]:
    root = root or GOOFISH_DATA_DIR
    summary_path = _resolve_summary_excel(root)
    wb = load_workbook(str(summary_path), read_only=True, data_only=True)
    try:
        ws = wb.active
        try:
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        except StopIteration:
            return []
        headers = [_cell_str(c.value) for c in header_cells]
        rows: List[Dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            item = {
                headers[i]: (_cell_str(row[i]) if i < len(row) else '')
                for i in range(len(headers))
            }
            rows.append(item)
        return rows
    finally:
        wb.close()


def load_pending_products(root: Optional[Path] = None) -> List[ProductRecord]:
    """有图且上架链接为空的商品。"""
    root = root or GOOFISH_DATA_DIR
    pending: List[ProductRecord] = []
    for row in _iter_summary_rows(root):
        if row.get(COL_UPLOAD_LINK):
            continue
        product_dir = _resolve_product_dir(row, root)
        if not product_dir:
            logger.warning('跳过无效目录: %s', row.get(COL_TITLE))
            continue
        product = _load_product_detail(product_dir, row)
        if product.upload_link:
            logger.info('待发布列表排除（单品已有链接）title=%s', product.title)
            continue
        if not product.images:
            logger.warning('跳过无图商品: %s', product.title)
            continue
        pending.append(product)
    return pending


def load_product_by_title(title: str, root: Optional[Path] = None) -> Optional[ProductRecord]:
    root = root or GOOFISH_DATA_DIR
    target = (title or '').strip()
    if not target:
        return None
    for row in _iter_summary_rows(root):
        if (row.get(COL_TITLE) or '').strip() == target:
            product_dir = _resolve_product_dir(row, root)
            if product_dir:
                return _load_product_detail(product_dir, row)
    return None


def load_product_by_keyword(keyword: str, root: Optional[Path] = None) -> Optional[ProductRecord]:
    root = root or GOOFISH_DATA_DIR
    kw = (keyword or '').strip()
    if not kw:
        return None
    for row in _iter_summary_rows(root):
        if kw in (row.get(COL_TITLE) or ''):
            product_dir = _resolve_product_dir(row, root)
            if product_dir:
                return _load_product_detail(product_dir, row)
    return None

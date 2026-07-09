"""上架成功后回填汇总表与单品 Excel。"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from spider.taobao.config import SUMMARY_EXCEL_NAME, TAOBAO_DATA_DIR
from utils.logger import get_logger

logger = get_logger('TaobaoBackfill')


def _safe_dirname(title: str) -> str:
    illegal = r'[\\/:*?"<>|]'
    name = re.sub(illegal, '_', title).strip()
    return name[:80]


def _resolve_summary_excel(root: Path) -> Path:
    named = root / SUMMARY_EXCEL_NAME
    if named.exists():
        return named
    for p in sorted(root.glob('*.xlsx')):
        if p.name.startswith('~$') or '.tmp' in p.name.lower():
            continue
        if '汇总' in p.name:
            return p
    return named


def sync_summary_upload_link(
    product,
    *,
    root: Optional[Path] = None,
) -> bool:
    """汇总表上架链接为空但单品已有链接时，将单品信息同步到汇总表。"""
    if not product.upload_link:
        return False
    root = root or TAOBAO_DATA_DIR
    return _backfill_summary(
        product.title,
        product.shop_name,
        product.upload_link,
        product.upload_time or datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        root,
    )


def backfill_upload_result(
    *,
    title: str,
    item_id: str,
    product_dir: Optional[Path] = None,
    shop_name: str = '',
    root: Optional[Path] = None,
) -> dict:
    """
    回填上架店铺/时间/链接到汇总 Excel 与单品 Excel。

    Returns:
        { ok, item_url, upload_time, summary_updated, product_updated }
    """
    root = root or TAOBAO_DATA_DIR
    upload_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    item_url = f'https://item.taobao.com/item.htm?id={item_id}'
    product_dir = product_dir or (root / _safe_dirname(title))

    summary_updated = _backfill_summary(title, shop_name, item_url, upload_time, root)
    product_updated = _backfill_product_excel(product_dir, shop_name, item_url, upload_time)

    logger.info('回填完成 title=%s item_id=%s', title, item_id)
    return {
        'ok': True,
        'item_url': item_url,
        'upload_time': upload_time,
        'summary_updated': summary_updated,
        'product_updated': product_updated,
    }


def _backfill_summary(title: str, shop_name: str, item_url: str, upload_time: str, root: Path) -> bool:
    summary_path = _resolve_summary_excel(root)
    if not summary_path.exists():
        return False
    wb = load_workbook(str(summary_path))
    ws = wb.active
    header_row = [c.value for c in ws[1]]
    try:
        col_shop = header_row.index('上架店铺') + 1
        col_time = header_row.index('上架时间') + 1
        col_link = header_row.index('上架链接') + 1
    except ValueError:
        return False

    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == title:
            ws.cell(row=row[0].row, column=col_shop, value=shop_name)
            ws.cell(row=row[0].row, column=col_time, value=upload_time)
            ws.cell(row=row[0].row, column=col_link, value=item_url)
            found = True
            break
    wb.save(str(summary_path))
    return found


def _backfill_product_excel(product_dir: Path, shop_name: str, item_url: str, upload_time: str) -> bool:
    excel_path = product_dir / '商品信息.xlsx'
    if not excel_path.exists():
        return False
    wb = load_workbook(str(excel_path))
    ws = wb['基本信息']
    upload_fields = {'上架店铺': shop_name, '上架时间': upload_time, '上架链接': item_url}
    existing = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    for field, value in upload_fields.items():
        if field in existing:
            ws.cell(row=existing[field], column=2, value=value)
        else:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=field)
            ws.cell(row=next_row, column=2, value=value)
    wb.save(str(excel_path))
    return True

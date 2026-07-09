"""从 Excel 与 images 目录加载待上架商品。"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from spider.taobao.config import SUMMARY_EXCEL_NAME, TAOBAO_DATA_DIR
from utils.logger import get_logger

logger = get_logger('TaobaoLoader')

_IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}


@dataclass
class ProductRecord:
    title: str
    price: str
    product_dir: Path
    product_url: str = ''
    images: List[Path] = field(default_factory=list)
    specs: List[Dict[str, Any]] = field(default_factory=list)
    skus: List[Dict[str, Any]] = field(default_factory=list)
    params: List[Dict[str, str]] = field(default_factory=list)
    shop_name: str = ''
    upload_link: str = ''
    upload_time: str = ''

    @property
    def slug(self) -> str:
        illegal = r'[\\/:*?"<>|]'
        name = re.sub(illegal, '_', self.title).strip()
        return name[:40] or 'product'

    @property
    def brand_short_name(self) -> str:
        """品名/型号：取标题前 30 字，去掉品牌常见前缀。"""
        t = self.title.strip()
        for sep in (' ', '　', '|', '/'):
            if sep in t:
                parts = [p.strip() for p in t.split(sep) if p.strip()]
                if parts:
                    return parts[0][:30]
        return t[:30]

    @property
    def spec_count(self) -> int:
        return sum(len(s.get('values', [])) for s in self.specs)

    def to_dict(self) -> Dict[str, Any]:
        return {
            'title': self.title,
            'price': self.price,
            'product_dir': str(self.product_dir),
            'product_url': self.product_url,
            'images': [str(p) for p in self.images],
            'specs': self.specs,
            'skus': self.skus,
            'params': self.params,
            'shop_name': self.shop_name,
            'upload_link': self.upload_link,
            'upload_time': self.upload_time,
            'spec_count': self.spec_count,
            'image_count': len(self.images),
        }


def _resolve_summary_excel(root: Path) -> Path:
    candidates = [
        p for p in sorted(root.glob('*.xlsx'))
        if not p.name.startswith('~$') and '.tmp' not in p.name.lower()
    ]
    for p in candidates:
        if p.name == SUMMARY_EXCEL_NAME or '汇总' in p.name:
            return p
    named = root / SUMMARY_EXCEL_NAME
    if named.exists():
        return named
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f'未找到汇总 Excel: {root / SUMMARY_EXCEL_NAME}')


def _cell_str(val: Any) -> str:
    if val is None:
        return ''
    return str(val).strip()


def _read_sheet_kv(ws) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out[_cell_str(row[0])] = _cell_str(row[1] if len(row) > 1 else '')
    return out


def _read_specs(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    specs_map: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        if not row or not row[0]:
            continue
        label = _cell_str(row[0])
        if label.startswith('（'):
            continue
        text = _cell_str(row[1] if len(row) > 1 else '')
        price = _cell_str(row[2] if len(row) > 2 else '')
        img = _cell_str(row[3] if len(row) > 3 else '')
        vid = _cell_str(row[4] if len(row) > 4 else '')
        empty = _cell_str(row[5] if len(row) > 5 else '') == '缺货'
        specs_map.setdefault(label, []).append({
            'text': text,
            'price': price,
            'img': img,
            'vid': vid,
            'empty': empty,
        })
    return [{'label': k, 'values': v} for k, v in specs_map.items()]


def _read_skus(ws) -> List[Dict[str, Any]]:
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        sku_id = _cell_str(row[0])
        if sku_id.startswith('（'):
            continue
        names_raw = _cell_str(row[1] if len(row) > 1 else '')
        names = [n.strip() for n in names_raw.split('/') if n.strip()]
        out.append({
            'skuId': sku_id,
            'names': names,
            'price': _cell_str(row[2] if len(row) > 2 else ''),
            'empty': _cell_str(row[3] if len(row) > 3 else '') == '缺货',
        })
    return out


def _read_params(ws) -> List[Dict[str, str]]:
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out.append({
            'label': _cell_str(row[0]),
            'value': _cell_str(row[1] if len(row) > 1 else ''),
            'type': _cell_str(row[2] if len(row) > 2 else ''),
        })
    return out


def _collect_images(product_dir: Path) -> List[Path]:
    img_dir = product_dir / 'images'
    if not img_dir.is_dir():
        return []
    files = [
        p for p in img_dir.iterdir()
        if p.is_file() and p.suffix.lower() in _IMAGE_EXTS
    ]
    return sorted(files, key=lambda p: p.name.lower())


def _load_product_detail(product_dir: Path, summary_row: Dict[str, str]) -> ProductRecord:
    excel_path = product_dir / '商品信息.xlsx'
    title = summary_row.get('商品标题') or product_dir.name
    price = summary_row.get('价格') or ''
    product_url = summary_row.get('商品URL') or ''
    upload_link = summary_row.get('上架链接') or ''
    upload_time = summary_row.get('上架时间') or ''
    shop_name = summary_row.get('上架店铺') or ''

    if excel_path.exists():
        wb = load_workbook(str(excel_path), read_only=True, data_only=True)
        try:
            basic = _read_sheet_kv(wb['基本信息']) if '基本信息' in wb.sheetnames else {}
            title = basic.get('商品标题') or title
            price = basic.get('价格（元）') or basic.get('价格') or price
            product_url = basic.get('商品URL') or product_url
            shop_name = basic.get('上架店铺') or shop_name
            upload_link = basic.get('上架链接') or upload_link
            upload_time = basic.get('上架时间') or upload_time
            specs = _read_specs(wb['规格']) if '规格' in wb.sheetnames else []
            skus = _read_skus(wb['SKU价格']) if 'SKU价格' in wb.sheetnames else []
            params = _read_params(wb['参数']) if '参数' in wb.sheetnames else []
        finally:
            wb.close()
    else:
        specs, skus, params = [], [], []

    images = _collect_images(product_dir)
    return ProductRecord(
        title=title,
        price=price,
        product_dir=product_dir,
        product_url=product_url,
        images=images,
        specs=specs,
        skus=skus,
        params=params,
        shop_name=shop_name,
        upload_link=upload_link,
        upload_time=upload_time,
    )


def _iter_summary_rows(root: Optional[Path] = None) -> List[Dict[str, str]]:
    root = root or TAOBAO_DATA_DIR
    summary_path = _resolve_summary_excel(root)
    wb = load_workbook(str(summary_path), read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = [_cell_str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows: List[Dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            item = {headers[i]: _cell_str(row[i]) if i < len(row) else '' for i in range(len(headers))}
            rows.append(item)
        return rows
    finally:
        wb.close()


def load_pending_products(root: Optional[Path] = None) -> List[ProductRecord]:
    """加载有图且上架链接为空的商品。"""
    root = root or TAOBAO_DATA_DIR
    pending: List[ProductRecord] = []
    for row in _iter_summary_rows(root):
        upload_link = row.get('上架链接') or ''
        if upload_link:
            continue
        product_dir = Path(row.get('本地目录') or '')
        if not product_dir.is_dir():
            product_dir = root / product_dir.name if product_dir.name else None
        if not product_dir or not product_dir.is_dir():
            logger.warning('跳过无效目录: %s', row.get('商品标题'))
            continue
        product = _load_product_detail(product_dir, row)
        if product.upload_link:
            logger.info(
                '待上架列表排除（单品已有链接）title=%s link=%s summary_link=(空)',
                product.title,
                product.upload_link,
            )
            continue
        if not product.images:
            logger.warning('跳过无图商品: %s', product.title)
            continue
        pending.append(product)
    return pending


def load_product_by_title(title: str, root: Optional[Path] = None) -> Optional[ProductRecord]:
    title = title.strip()
    for row in _iter_summary_rows(root):
        if row.get('商品标题') == title:
            product_dir = Path(row.get('本地目录') or '')
            if not product_dir.is_dir():
                root = root or TAOBAO_DATA_DIR
                product_dir = root / product_dir.name
            if product_dir.is_dir():
                return _load_product_detail(product_dir, row)
    return None


def load_product_by_keyword(keyword: str, root: Optional[Path] = None) -> Optional[ProductRecord]:
    keyword = keyword.strip()
    if not keyword:
        return None
    for row in _iter_summary_rows(root):
        title = row.get('商品标题') or ''
        if keyword in title:
            product_dir = Path(row.get('本地目录') or '')
            if not product_dir.is_dir():
                root = root or TAOBAO_DATA_DIR
                product_dir = root / product_dir.name
            if product_dir.is_dir():
                return _load_product_detail(product_dir, row)
    return None

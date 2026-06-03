"""
淘宝商品数据保存 API

POST /api/taobao/save-product
  - 接收 webAuto popup 发来的商品数据
  - 下载图片到本地
  - 写入单品 Excel（基本信息 / 规格 / 参数 / 图片 四个 Sheet）
  - 更新根目录汇总 Excel 与 README.md
"""
import os
import re
import json
import requests
from datetime import datetime
from pathlib import Path

from flask import Blueprint, request, jsonify
from flasgger import swag_from
from utils.logger import get_logger

logger = get_logger('TaobaoRoutes')
bp = Blueprint('taobao', __name__, url_prefix='/api/taobao')

# 保存根目录
SAVE_ROOT = Path(r'C:\Users\yao\Desktop\work\电商数据\淘宝')
SUMMARY_EXCEL = SAVE_ROOT / '淘宝商品汇总.xlsx'
SUMMARY_MD = SAVE_ROOT / 'README.md'

# 汇总 Excel 列头
SUMMARY_HEADERS = [
    '商品标题', '价格', '已售', '规格数', '参数数', '图片数',
    '本地目录', '商品URL', '采集时间',
    '上架店铺', '上架时间', '上架链接',
]


def _safe_dirname(title: str) -> str:
    """将商品标题转换为安全的文件夹名（去掉 Windows 非法字符，截断到 80 字符）"""
    illegal = r'[\\/:*?"<>|]'
    name = re.sub(illegal, '_', title).strip()
    return name[:80]


def _download_images(image_urls: list, img_dir: Path) -> list:
    """下载图片列表到 img_dir，返回 [{ url, filename, ok }]"""
    img_dir.mkdir(parents=True, exist_ok=True)
    results = []
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Referer': 'https://item.taobao.com/',
    }
    for idx, url in enumerate(image_urls, start=1):
        # 从 URL 推断扩展名
        ext = '.jpg'
        for candidate in ['.jpg', '.jpeg', '.png', '.webp', '.gif']:
            if candidate in url.lower():
                ext = candidate
                break
        filename = f'{idx:02d}{ext}'
        filepath = img_dir / filename
        try:
            resp = requests.get(url, headers=headers, timeout=15)
            resp.raise_for_status()
            filepath.write_bytes(resp.content)
            results.append({'url': url, 'filename': filename, 'ok': True})
            logger.info(f'[图片下载] {filename} ← {url[:60]}…')
        except Exception as e:
            results.append({'url': url, 'filename': filename, 'ok': False, 'error': str(e)})
            logger.warning(f'[图片下载] 失败 {url[:60]}… | {e}')
    return results


def _write_product_excel(product_dir: Path, data: dict, img_results: list):
    """写单品 Excel（4 个 Sheet）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── 样式工具 ────────────────────────────────────────────────
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill('solid', fgColor='2F75B6')
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    def style_cell(cell, align=LEFT):
        cell.alignment = align
        cell.border = BORDER

    def auto_width(ws, min_w=10, max_w=60):
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max((len(str(c.value or '')) for c in col_cells), default=0)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # ── Sheet 1：基本信息 ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = '基本信息'
    ws1.row_dimensions[1].height = 22
    headers1 = ['字段', '值']
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        style_header(c)
    rows1 = [
        ('商品标题', data.get('title', '')),
        ('价格（元）', data.get('price', '')),
        ('已售数量', data.get('sold', '')),
        ('商品URL', data.get('url', '')),
        ('采集时间', now_str),
    ]
    for r, (field, value) in enumerate(rows1, 2):
        ws1.cell(row=r, column=1, value=field).border = BORDER
        ws1.cell(row=r, column=2, value=value).border = BORDER
    auto_width(ws1)

    # ── Sheet 2：规格 ───────────────────────────────────────────
    ws2 = wb.create_sheet('规格')
    headers2 = ['规格标签', '规格值', '价格', '图片URL', 'vid', '缺货']
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        style_header(c)
    row2 = 2
    for spec in data.get('specs', []):
        label = spec.get('label', '')
        for val in spec.get('values', []):
            # 兼容旧格式（字符串）和新格式（对象）
            if isinstance(val, dict):
                text  = val.get('text', '')
                price = val.get('price') or ''
                img   = val.get('img') or ''
                vid   = val.get('vid') or ''
                empty = '缺货' if val.get('empty') else ''
            else:
                text = str(val)
                price = img = vid = empty = ''
            ws2.cell(row=row2, column=1, value=label).border = BORDER
            ws2.cell(row=row2, column=2, value=text).border  = BORDER
            ws2.cell(row=row2, column=3, value=price).border = BORDER
            ws2.cell(row=row2, column=4, value=img).border   = BORDER
            ws2.cell(row=row2, column=5, value=vid).border   = BORDER
            ws2.cell(row=row2, column=6, value=empty).border = BORDER
            row2 += 1
    if row2 == 2:
        ws2.cell(row=2, column=1, value='（无规格数据）').border = BORDER
    auto_width(ws2)

    # ── Sheet 3：SKU价格 ─────────────────────────────────────────
    ws_sku = wb.create_sheet('SKU价格')
    headers_sku = ['skuId', '规格组合', '价格', '缺货']
    for col, h in enumerate(headers_sku, 1):
        c = ws_sku.cell(row=1, column=col, value=h)
        style_header(c)
    skus = data.get('skus', [])
    for r, sku in enumerate(skus, 2):
        ws_sku.cell(row=r, column=1, value=sku.get('skuId') or '').border  = BORDER
        ws_sku.cell(row=r, column=2, value=' / '.join(sku.get('names', []))).border = BORDER
        ws_sku.cell(row=r, column=3, value=sku.get('price') or '').border  = BORDER
        ws_sku.cell(row=r, column=4, value='缺货' if sku.get('empty') else '').border = BORDER
    if not skus:
        ws_sku.cell(row=2, column=1, value='（无SKU数据）').border = BORDER
    auto_width(ws_sku)

    # ── Sheet 4：参数 ───────────────────────────────────────────
    ws3 = wb.create_sheet('参数')
    headers3 = ['参数名', '参数值', '类型']
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        style_header(c)
    for r, param in enumerate(data.get('params', []), 2):
        ws3.cell(row=r, column=1, value=param.get('label', '')).border = BORDER
        ws3.cell(row=r, column=2, value=param.get('value', '')).border = BORDER
        ws3.cell(row=r, column=3, value=param.get('type', '')).border = BORDER
    if not data.get('params'):
        ws3.cell(row=2, column=1, value='（无参数数据）').border = BORDER
    auto_width(ws3)

    # ── Sheet 5：图片 ───────────────────────────────────────────
    ws4 = wb.create_sheet('图片')
    headers4 = ['序号', '原始URL', '本地文件名', '下载状态']
    for col, h in enumerate(headers4, 1):
        c = ws4.cell(row=1, column=col, value=h)
        style_header(c)
    for r, img in enumerate(img_results, 2):
        ws4.cell(row=r, column=1, value=r - 1).border = BORDER
        ws4.cell(row=r, column=2, value=img.get('url', '')).border = BORDER
        ws4.cell(row=r, column=3, value=img.get('filename', '')).border = BORDER
        ws4.cell(row=r, column=4, value='✓ 成功' if img.get('ok') else f'✗ 失败: {img.get("error", "")}').border = BORDER
    if not img_results:
        ws4.cell(row=2, column=1, value='（无图片数据）').border = BORDER
    auto_width(ws4)

    wb.save(str(product_dir / '商品信息.xlsx'))
    logger.info(f'[Excel] 写入 {product_dir / "商品信息.xlsx"}')


def _update_summary_excel(data: dict, product_dir: Path, img_count: int):
    """追加/更新汇总 Excel（若标题已存在则更新那一行，否则追加）"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill('solid', fgColor='2F75B6')
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = data.get('title', '')
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    new_row = [
        title,
        data.get('price', ''),
        data.get('sold', ''),
        sum(len(s.get('values', [])) for s in data.get('specs', [])),
        len(data.get('params', [])),
        img_count,
        str(product_dir),
        data.get('url', ''),
        now_str,
        data.get('shopName', ''),   # 上架店铺（save时为空，mark-uploaded时回填）
        data.get('uploadTime', ''), # 上架时间
        data.get('itemUrl', ''),    # 上架链接
    ]

    if SUMMARY_EXCEL.exists():
        wb = load_workbook(str(SUMMARY_EXCEL))
        ws = wb.active
        # 找到标题列（第1列），看是否已有该商品
        found_row = None
        for row in ws.iter_rows(min_row=2):
            if row[0].value == title:
                found_row = row[0].row
                break
        if found_row:
            for col_idx, val in enumerate(new_row, 1):
                ws.cell(row=found_row, column=col_idx, value=val)
            logger.info(f'[汇总Excel] 更新第 {found_row} 行: {title}')
        else:
            ws.append(new_row)
            logger.info(f'[汇总Excel] 追加新行: {title}')
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '淘宝商品汇总'
        ws.row_dimensions[1].height = 22
        for col, h in enumerate(SUMMARY_HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
        ws.append(new_row)
        logger.info(f'[汇总Excel] 新建文件并写入: {title}')

    # 自动列宽
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value or '')) for c in col_cells), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 80)

    wb.save(str(SUMMARY_EXCEL))


def _update_summary_md(data: dict, product_dir: Path, img_count: int):
    """更新汇总 README.md（若标题已存在则覆盖那一行，否则追加）"""
    title = data.get('title', '')
    price = data.get('price', '')
    sold = data.get('sold', '')
    url = data.get('url', '')
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    spec_count = sum(len(s.get('values', [])) for s in data.get('specs', []))
    dir_name = product_dir.name

    # 新行内容（含上架信息占位列）
    new_line = (
        f'| [{title}](./{dir_name}/) '
        f'| {price} '
        f'| {sold} '
        f'| {spec_count} '
        f'| {len(data.get("params", []))} '
        f'| {img_count} '
        f'| {now_str} '
        f'| [链接]({url}) '
        f'|  '   # 上架店铺（待回填）
        f'|  '   # 上架时间（待回填）
        f'|  |'  # 上架链接（待回填）
    )

    md_header = (
        '# 淘宝商品数据汇总\n\n'
        '> 由 webAuto 淘宝助手自动生成，每次保存商品时追加/更新。\n\n'
        '| 商品标题 | 价格 | 已售 | 规格值数 | 参数数 | 图片数 | 采集时间 | 商品链接 | 上架店铺 | 上架时间 | 上架链接 |\n'
        '|----------|------|------|---------|--------|--------|---------|----------|---------|---------|----------|\n'
    )

    if SUMMARY_MD.exists():
        content = SUMMARY_MD.read_text(encoding='utf-8')
        lines = content.split('\n')
        # 找到含该标题的行并替换
        replaced = False
        new_lines = []
        for line in lines:
            if f'](./{dir_name}/)' in line or (f'[{title}]' in line):
                new_lines.append(new_line)
                replaced = True
            else:
                new_lines.append(line)
        if not replaced:
            # 在最后一个 | 行之后插入
            new_lines.append(new_line)
        SUMMARY_MD.write_text('\n'.join(new_lines), encoding='utf-8')
        logger.info(f'[汇总MD] {"更新" if replaced else "追加"}: {title}')
    else:
        SUMMARY_MD.write_text(md_header + new_line + '\n', encoding='utf-8')
        logger.info(f'[汇总MD] 新建文件并写入: {title}')


@bp.route('/save-product', methods=['POST'])
@swag_from({
    'tags': ['淘宝'],
    'summary': '保存淘宝商品数据到本地',
    'parameters': [{
        'in': 'body',
        'name': 'body',
        'schema': {
            'type': 'object',
            'properties': {
                'title': {'type': 'string', 'description': '商品标题'},
                'price': {'type': 'string', 'description': '价格'},
                'sold': {'type': 'string', 'description': '已售数量'},
                'url': {'type': 'string', 'description': '商品页面URL'},
                'images': {'type': 'array', 'items': {'type': 'string'}, 'description': '图片URL列表'},
                'specs': {'type': 'array', 'description': '规格列表（values 为对象数组，含 text/img/vid/empty/price）'},
                'skus':  {'type': 'array', 'description': 'SKU价格列表（含 skuId/vids/names/price/empty）'},
                'params': {'type': 'array', 'description': '参数列表'},
            },
            'required': ['title'],
        }
    }],
    'responses': {
        200: {'description': '保存成功'},
        400: {'description': '参数错误'},
        500: {'description': '服务器错误'},
    }
})
def taobao_save_product():
    """保存淘宝商品数据（下载图片 + 写单品Excel + 更新汇总Excel/MD）"""
    try:
        body = request.get_json(force=True) or {}
        title = (body.get('title') or '').strip()
        if not title:
            return jsonify({'ok': False, 'error': '商品标题不能为空'}), 400

        # 准备目录
        SAVE_ROOT.mkdir(parents=True, exist_ok=True)
        dir_name = _safe_dirname(title)
        product_dir = SAVE_ROOT / dir_name
        img_dir = product_dir / 'images'
        product_dir.mkdir(parents=True, exist_ok=True)

        logger.info(f'[保存商品] 标题: {title}')
        logger.info(f'[保存商品] 目录: {product_dir}')

        # 1. 下载图片
        image_urls = body.get('images') or []
        img_results = _download_images(image_urls, img_dir) if image_urls else []
        img_ok = sum(1 for r in img_results if r.get('ok'))

        # 2. 写单品 Excel
        _write_product_excel(product_dir, body, img_results)

        # 3. 更新汇总
        _update_summary_excel(body, product_dir, len(img_results))
        _update_summary_md(body, product_dir, len(img_results))

        return jsonify({
            'ok': True,
            'folder': str(product_dir),
            'imageTotal': len(img_results),
            'imageOk': img_ok,
            'log': [
                f'✓ 商品目录: {product_dir}',
                f'✓ 图片下载: {img_ok}/{len(img_results)} 张成功',
                '✓ 单品 Excel 已写入',
                '✓ 汇总 Excel 已更新',
                '✓ 汇总 README.md 已更新',
            ]
        })

    except Exception as e:
        logger.error(f'[保存商品] 异常: {e}', exc_info=True)
        return jsonify({'ok': False, 'error': str(e)}), 500


def _backfill_summary_excel(title: str, shop_name: str, item_url: str, upload_time: str) -> bool:
    """在汇总 Excel 中找到 title 对应行，将上架信息回填到最后三列，返回是否找到并更新。"""
    if not SUMMARY_EXCEL.exists():
        return False
    from openpyxl import load_workbook
    wb = load_workbook(str(SUMMARY_EXCEL))
    ws = wb.active

    # 找 '上架店铺' 等列索引（兼容列序）
    header_row = [c.value for c in ws[1]]
    try:
        col_shop   = header_row.index('上架店铺')   + 1
        col_time   = header_row.index('上架时间')   + 1
        col_link   = header_row.index('上架链接')   + 1
    except ValueError:
        # 列不存在则追加列头
        next_col = len(header_row) + 1
        ws.cell(row=1, column=next_col,     value='上架店铺')
        ws.cell(row=1, column=next_col + 1, value='上架时间')
        ws.cell(row=1, column=next_col + 2, value='上架链接')
        col_shop = next_col
        col_time = next_col + 1
        col_link = next_col + 2

    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == title:
            ws.cell(row=row[0].row, column=col_shop, value=shop_name)
            ws.cell(row=row[0].row, column=col_time, value=upload_time)
            ws.cell(row=row[0].row, column=col_link, value=item_url)
            found = True
            logger.info(f'[汇总Excel] 回填上架信息 第{row[0].row}行: {title}')
            break

    wb.save(str(SUMMARY_EXCEL))
    return found


def _backfill_product_excel(product_dir: Path, shop_name: str, item_url: str, upload_time: str) -> bool:
    """在单品 Excel 基本信息 Sheet 追加/更新上架信息行。"""
    excel_path = product_dir / '商品信息.xlsx'
    if not excel_path.exists():
        return False
    from openpyxl import load_workbook
    wb = load_workbook(str(excel_path))
    ws = wb['基本信息']

    # 找已有的上架相关行，若存在则更新，否则追加
    upload_fields = {'上架店铺': shop_name, '上架时间': upload_time, '上架链接': item_url}
    existing_labels = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    for field, value in upload_fields.items():
        if field in existing_labels:
            ws.cell(row=existing_labels[field], column=2, value=value)
        else:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=field)
            ws.cell(row=next_row, column=2, value=value)

    wb.save(str(excel_path))
    return True


@bp.route('/mark-uploaded', methods=['POST'])
@swag_from({
    'tags': ['淘宝'],
    'summary': '商品上架后回填店铺信息到汇总表与单品Excel',
    'parameters': [{
        'in': 'body',
        'name': 'body',
        'schema': {
            'type': 'object',
            'properties': {
                'title':      {'type': 'string', 'description': '商品标题（用于匹配汇总表行）'},
                'shopName':   {'type': 'string', 'description': '上架店铺名称'},
                'itemUrl':    {'type': 'string', 'description': '上架后商品链接'},
                'uploadTime': {'type': 'string', 'description': '上架时间（默认当前时间）'},
            },
            'required': ['title'],
        }
    }],
    'responses': {
        200: {'description': '回填成功'},
        404: {'description': '汇总表中未找到该商品'},
        500: {'description': '服务器错误'},
    }
})
def taobao_mark_uploaded():
    """商品上架完成后，将店铺名/上架时间/商品链接回填到汇总表与单品Excel"""
    try:
        body = request.get_json(force=True) or {}
        title = (body.get('title') or '').strip()
        if not title:
            return jsonify({'ok': False, 'error': '商品标题不能为空'}), 400

        shop_name   = (body.get('shopName')   or '').strip()
        item_url    = (body.get('itemUrl')    or '').strip()
        upload_time = (body.get('uploadTime') or datetime.now().strftime('%Y-%m-%d %H:%M:%S')).strip()

        log = []

        # 1. 回填汇总 Excel
        found = _backfill_summary_excel(title, shop_name, item_url, upload_time)
        if found:
            log.append('✓ 汇总 Excel 已回填上架信息')
        else:
            log.append('⚠ 汇总 Excel 未找到该商品标题，跳过')

        # 2. 回填单品 Excel（若目录存在）
        dir_name = _safe_dirname(title)
        product_dir = SAVE_ROOT / dir_name
        if product_dir.exists():
            ok2 = _backfill_product_excel(product_dir, shop_name, item_url, upload_time)
            log.append('✓ 单品 Excel 基本信息已更新' if ok2 else '⚠ 单品 Excel 未找到，跳过')
        else:
            log.append('⚠ 单品目录不存在，跳过单品 Excel 更新')

        # 3. 更新汇总 README.md（找到对应行追加上架信息）
        if SUMMARY_MD.exists():
            content = SUMMARY_MD.read_text(encoding='utf-8')
            dir_link = f'](./{dir_name}/)'
            new_lines = []
            replaced = False
            for line in content.split('\n'):
                if dir_link in line or (f'[{title}]' in line and '|' in line):
                    # 若行末没有上架信息列，追加；否则更新末尾三列
                    parts = line.rstrip('|').split('|')
                    # 确保至少有 9 个数据列（标题到采集时间）
                    while len(parts) < 10:
                        parts.append(' ')
                    # 将最后三列替换为上架信息
                    parts = parts[:10] + [f' {shop_name} ', f' {upload_time} ', f' {item_url} ']
                    line = '|'.join(parts) + '|'
                    replaced = True
                new_lines.append(line)
            SUMMARY_MD.write_text('\n'.join(new_lines), encoding='utf-8')
            log.append('✓ 汇总 README.md 已更新' if replaced else '⚠ README.md 未匹配到商品行，跳过')

        return jsonify({
            'ok': True,
            'found': found,
            'log': log,
        })

    except Exception as e:
        logger.error(f'[mark-uploaded] 异常: {e}', exc_info=True)
        return jsonify({'ok': False, 'error': str(e)}), 500
